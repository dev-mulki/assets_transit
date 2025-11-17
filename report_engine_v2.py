import os, io, openpyxl, datetime, tempfile, subprocess
from flask import send_file, current_app
from sqlalchemy import text
from copy import copy
from openpyxl.styles import Font, Alignment, Border, Side
from app import db


# ================================================================
#  HELPER: SAFE FORMAT QUERY
# ================================================================
def safe_format(q, p):
    class SafeDict(dict):
        def __missing__(self, key):
            return "NULL"

        def __getitem__(self, key):
            val = dict.get(self, key)
            if val is None or str(val).lower() in ["none", "null", "undefined", ""]:
                return "NULL"
            if isinstance(val, (int, float)):
                return str(val)
            safe_val = str(val).replace("'", "''")
            return f"'{safe_val}'"

    return q.format_map(SafeDict(p))


# ================================================================
#  HELPER: STYLE
# ================================================================
thin = Side(border_style="thin", color="000000")
border_default = Border(top=thin, left=thin, right=thin, bottom=thin)


def extract_sign_block(ws):
    sign = []

    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if not isinstance(cell.value, str):
                continue

            raw_val = cell.value

            # DETEKSI _sign DI DALAM PLACEHOLDER {}, contoh: {nama_unit_sign}
            if "_sign" in raw_val:
                # simpan posisi + placeholder asli (BELUM format)
                sign.append((
                    cell.row,
                    cell.column,
                    raw_val,      # simpan teks asli
                    copy(cell.font),
                    copy(cell.alignment)
                ))

                # hapus cell supaya tidak ketiban subreport
                cell.value = None

    return sign



def get_alignment(align, wrap=False):
    align = (align or "left").lower()
    return Alignment(
        horizontal=align if align in ["left", "center", "right"] else "left",
        wrapText=wrap,
        vertical="center",
        indent=1 if align == "left" else 0
    )


def apply_style(cell, row, style_rule):
    if not style_rule:
        return
    rules = [r.strip() for r in style_rule.split(",")]

    for rule in rules:
        try:
            if rule == "bold":
                cell.font = Font(bold=True)
            elif rule == "italic":
                cell.font = Font(italic=True)
            elif rule == "underline":
                cell.font = Font(underline="single")

            elif rule.startswith("bold_if:"):
                cond = rule.split(":", 1)[1]
                if eval(str(row.get("level", 0)) + cond.replace("level", "")):
                    cell.font = Font(bold=True)

            elif rule.startswith("red_if:"):
                cond = rule.split(":", 1)[1]
                if eval(str(cell.value) + cond):
                    cell.font = Font(color="FF0000")

            elif rule.startswith("green_if:"):
                cond = rule.split(":", 1)[1]
                if eval(str(cell.value) + cond):
                    cell.font = Font(color="00AA00")

        except:
            continue


# ================================================================
#  HELPER: DETEKSI MARKER ~~subreport.xlsx~~
# ================================================================
def extract_subreport_markers(sheet_output):
    markers = []

    for row in sheet_output.iter_rows(values_only=False):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("~~") and cell.value.endswith("~~"):
                filename = cell.value.strip("~")
                markers.append((cell.row, cell.column, filename))
                cell.value = None  # Hapus placeholder

    return markers


# ================================================================
#  CORE ENGINE: PROCESS SINGLE REPORT (TANPA SUBREPORT)
# ================================================================
def process_single_report(file_path, params):
    wb = openpyxl.load_workbook(file_path)
    sheet_output = wb["Output"]
    sheet_query = wb["Query"]
    sheet_mapping = wb["Mapping"]

    # --------------------------
    # READ QUERY
    # --------------------------
    query_parts, row_idx = [], 3
    while True:
        val = sheet_query.cell(row=row_idx, column=2).value
        if not val:
            break
        query_parts.append(str(val).strip())
        row_idx += 1

    query = safe_format(" ".join(query_parts), params)
    result = db.session.execute(text(query))
    columns = result.keys()
    data = [dict(zip(columns, row)) for row in result.fetchall()]

    # --------------------------
    # READ MAPPING
    # --------------------------
    start_row, start_col = 1, 1
    if sheet_mapping["A1"].value and "start_row_from" in str(sheet_mapping["A1"].value).lower():
        start_row = int(sheet_mapping["B1"].value or 1)
    if sheet_mapping["A2"].value and "start_col_from" in str(sheet_mapping["A2"].value).lower():
        start_col = int(sheet_mapping["B2"].value or 1)

    mapping = []
    for row in sheet_mapping.iter_rows(min_row=4, values_only=True):
        if not row[0]:
            continue
        sheet_col, db_col, dtype, prefix, suffix, align, style, aggregate, wraptext = row + (None,) * (9 - len(row))
        mapping.append({
            "sheet_col": sheet_col,
            "db_col": db_col,
            "dtype": (dtype or "string").lower(),
            "prefix": prefix or "",
            "suffix": suffix or "",
            "align": (align or "left").lower(),
            "style": style or "",
            "aggregate": str(aggregate).lower() if aggregate else "",
            "wraptext": str(wraptext).strip().lower() in ["1", "true", "yes", "y"],
        })

    # --------------------------
    # DETEKSI SUBREPORT
    # --------------------------
    subreport_markers = extract_subreport_markers(sheet_output)

    print(subreport_markers)
    # ---- Geser tanda tangan (_sign) ----
    # --- Deteksi blok tanda tangan di template
    # --- Deteksi semua cell tanda tangan

    # --------------------------
    # REPLACE {} PLACEHOLDER (NO SUBREPORT)
    # --------------------------
    merged_vars = {**params, **(data[0] if data else {})}

    for row in sheet_output.iter_rows(values_only=False):
        for cell in row:
            if not cell.value or not isinstance(cell.value, str):
                continue

            # 🚨 JANGAN format cell yang mengandung _sign
            if "_sign" in cell.value:
                print(cell.value)
                pass

            if "{" in cell.value:
                try:
                    cell.value = cell.value.format(**merged_vars)
                except:
                    pass

    # --------------------------
    # WRITE MAIN DATA
    # --------------------------
    for i, row in enumerate(data):
        for j, m in enumerate(mapping, start=1):
            val = row.get(m["db_col"])
            dtype = m["dtype"]

            try:
                if dtype == "currency":
                    val = float(val or 0)
                    val = f"{m['prefix']}{val:,.0f}{m['suffix']}"
                elif dtype == "percent":
                    val = float(val or 0)
                    val = f"{val:.2f}{m['suffix']}"
                elif dtype in ["number", "int"]:
                    val = int(val or 0)
                elif dtype in ["float", "decimal"]:
                    val = float(val or 0)
                elif dtype == "date":
                    if isinstance(val, (datetime.datetime, datetime.date)):
                        val = val.strftime("%d-%m-%Y")
                elif dtype == "datetime":
                    if isinstance(val, (datetime.datetime, datetime.date)):
                        val = val.strftime("%d-%m-%Y %H:%M:%S")
                elif dtype == "boolean":
                    val = "Ya" if str(val).lower() in ["1", "true", "yes", "y"] else "Tidak"
            except:
                val = val or ""

            c = sheet_output.cell(row=start_row + i, column=start_col + j - 1, value=val)
            c.alignment = get_alignment(m["align"], m["wraptext"])
            c.border = border_default
            apply_style(c, row, m.get("style"))

    # --------------------------
    # AGGREGATE ROW
    # --------------------------
    if data and any(m.get("aggregate") for m in mapping):
        total_row_idx = start_row + len(data)
        sheet_output.cell(row=total_row_idx, column=start_col, value="TOTAL").font = Font(bold=True)

        for j, m in enumerate(mapping, start=1):
            agg = (m.get("aggregate") or "").strip().lower()
            if not agg:
                continue

            try:
                values = []
                for r in data:
                    v = r.get(m["db_col"])
                    if v is None:
                        continue
                    try:
                        v = float(str(v).replace(",", "").replace("Rp", "").replace("%", ""))
                        values.append(v)
                    except:
                        continue

                agg_val = None
                if agg == "sum":
                    agg_val = sum(values)
                elif agg == "avg":
                    agg_val = sum(values) / len(values) if values else 0
                elif agg == "min":
                    agg_val = min(values) if values else 0
                elif agg == "max":
                    agg_val = max(values) if values else 0
                elif agg == "count":
                    agg_val = len(values)

                if agg_val is not None:
                    c = sheet_output.cell(row=total_row_idx, column=start_col + j - 1)
                    dtype = m["dtype"]

                    if dtype == "currency":
                        c.value = f"{agg_val:,.0f}"
                    elif dtype == "percent":
                        c.value = f"{agg_val:.2f}"
                    elif dtype in ["float", "decimal"]:
                        c.value = f"{agg_val:,.2f}"
                    else:
                        c.value = agg_val

                    c.font = Font(bold=True)
                    c.alignment = get_alignment(m["align"], m["wraptext"])
                    c.border = border_default

            except:
                continue

        # === BORDER TOTAL (FIX LINE YANG HILANG DI ENGINE V3) ===
        last_col = start_col + len(mapping) - 1
        for col in range(start_col, last_col + 1):
            cell = sheet_output.cell(row=total_row_idx, column=col)
            cell.border = Border(
                top=Side(style="thin", color="000000"),
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000")
            )

    total_rows_used = start_row + len(data)

    # --------------------------
    # HANDLE _SIGN BLOCK FIRST
    # (FINAL SHIFT AKAN DI generate_full_report)
    # --------------------------

    return {
        "wb": wb,
        "ws": sheet_output,
        "subreports": subreport_markers,
        "total_rows": total_rows_used,
        "start_row": start_row,
        "start_col": start_col,
        "data": data
    }


# ================================================================
#  PASTE SUBREPORT
# ================================================================
from openpyxl.utils import get_column_letter


def merge_range_string(start_row, start_col, end_row, end_col):
    return f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"


def paste_sheet(main_ws, sub_ws, start_row, start_col=1):
    max_row = sub_ws.max_row
    max_col = sub_ws.max_column

    # Copy cell value + style
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            src = sub_ws.cell(row=r, column=c)
            dest = main_ws.cell(row=start_row + r - 1, column=start_col + c - 1)

            dest.value = src.value

            if src.has_style:
                dest.font = copy(src.font)
                dest.alignment = copy(src.alignment)
                dest.border = copy(src.border)
                dest.fill = copy(src.fill)

    # Copy merged cells (FIX)
    for merged in sub_ws.merged_cells.ranges:
        top = start_row + merged.min_row - 1
        left = merged.min_col
        bottom = start_row + merged.max_row - 1
        right = merged.max_col

        rng = merge_range_string(top, left, bottom, right)

        main_ws.merge_cells(rng)

    return max_row


# ================================================================
#  CALCULATE OFFSET FOR SUBREPORT
# ================================================================
def calculate_offset(main_start_row, total_main_rows, marker_row, total_sub_rows):
    # hitung jarak di template
    gap = marker_row - (main_start_row + 1)
    # posisi paste subreport yang benar
    paste_row = total_main_rows + gap + total_sub_rows + 2

    return paste_row


# ================================================================
#  SHIFT SIGN AFTER MERGE
# ================================================================
def shift_sign_after_merge(ws, start_row, total_main_rows, total_sub_rows):
    if total_sub_rows == 0:
        return  # Jangan geser lagi

    sign_cells = [
        (cell.row, cell.column, cell.value, cell)
        for row in ws.iter_rows(values_only=False)
        for cell in row
        if isinstance(cell.value, str) and "_sign" in cell.value
    ]

    if not sign_cells:
        return

    min_sign = min(r for r, _, _, _ in sign_cells)
    max_sign = max(r for r, _, _, _ in sign_cells)
    block_height = max_sign - min_sign + 1

    gap_sign = min_sign - (start_row + 1)

    new_start = start_row + total_main_rows + total_sub_rows + gap_sign + 2

    for r, c, val, old_cell in sign_cells:
        offset = r - min_sign
        new_row = new_start + offset
        clean_val = val.replace("_sign", "")

        new_cell = ws.cell(row=new_row, column=c, value=clean_val)

        if old_cell.has_style:
            new_cell.font = copy(old_cell.font)
            new_cell.alignment = copy(old_cell.alignment)

        old_cell.value = None

def place_sign_block(ws, sign_block, target_row, params):
    if not sign_block:
        return

    min_row = min(r for r, c, v, f, a in sign_block)

    for r, c, v, font, align in sign_block:
        offset = r - min_row
        row_final = target_row + offset - 6

        # bersihkan _sign → contoh: {nama_unit_sign} → {nama_unit}
        clean_key = v.replace("_sign", "")

        # format placeholder
        try:
            formatted_val = clean_key.format(**params)
        except:
            formatted_val = clean_key

        # tuliskan ke cell
        cell = ws.cell(row=row_final, column=c, value=formatted_val)
        cell.font = font
        cell.alignment = align



# ================================================================
#  ENGINE V3: GENERATE FULL REPORT OTOMATIS
# ================================================================
def generate_full_report(main_file, params):
    folder = "app/static/format_report"
    main_path = os.path.join(folder, main_file)
    # 1️⃣ Load main workbook dahulu
    wb = openpyxl.load_workbook(main_path)
    main_ws = wb["Output"]

    # 2️⃣ EXTRACT SIGN BLOCK sebelum apapun ditulis
    sign_block = extract_sign_block(main_ws)
    main = process_single_report(main_path, params)

    # MERGE PARAMS + ROW DATA UTAMA (buat replace sign)
    merged_sign_params = dict(params)
    if main.get("data"):
        try:
            merged_sign_params.update(main["data"][0])
        except:
            pass

    main_ws = main["ws"]
    main_wb = main["wb"]
    total_main_rows = main["total_rows"]
    start_row_main = main["start_row"]


    print(sign_block)

    markers = main["subreports"]
    total_sub_rows = 0

    # LOOP SUBREPORT
    for (marker_row, marker_col, subfile) in markers:
        sub_path = os.path.join(folder, subfile)
        sub = process_single_report(sub_path, params)

        paste_row = calculate_offset(
            start_row_main,
            total_main_rows,
            marker_row,
            total_sub_rows
        )

        used_rows = paste_sheet(main_ws, sub["ws"], paste_row, start_col=1)
        total_sub_rows += used_rows

    # ⬇⬇⬇ TARUH SIGN KEMBALI DI BAWAH SUBREPORTS ⬇⬇⬇
    final_sign_row = start_row_main + total_main_rows + total_sub_rows + 2
    place_sign_block(main_ws, sign_block, final_sign_row, merged_sign_params)


    # FINAL SHIFT SIGN
    shift_sign_after_merge(
        main_ws,
        start_row_main,
        total_main_rows,
        total_sub_rows
    )

    for name in main_wb.sheetnames:
        if name != "Output":
            ws = main_wb[name]
            main_wb.remove(ws)

    return main_wb
